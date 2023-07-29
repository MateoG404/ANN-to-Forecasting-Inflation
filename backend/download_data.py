import os
import pandas as pd
import requests
import sys
import openpyxl 

class download_data:

    def __init__(self) :

        self. meses = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
        self.años = ["19", "20", "21", "22"]
        self._url_dane = "https://www.dane.gov.co/files/investigaciones/boletines/ipc/anexo_ipc_"
        self.current_directory = os.getcwd().replace("\\","/")
        self.excel_inflation_folder = os.path.join(os.getcwd(), "data", "excel_inflation")    

    # Función que extrae los datos de determinado rango de filas y columnas de todos los exceles de una carpeta y da como output un dataframe
    def extract_data(self, name_df, row_min, row_max, year, sheet, folder_path=None):
        if folder_path is None:
            folder_path = self.excel_inflation_folder    
        
        dataset = []

        # recorremos todos los archivos dentro de la ruta indicada
        for filename in os.listdir(folder_path):
            # seleccionamos los archivos que terminen con el año seleccionado y que sean de tipo .xlxs
            if filename.endswith(str(year)+'.xlsx'):

                # Creamos la ruta del archivo uniendo la ruta del directorio y el nombre del archivo    
                file_path = os.path.join(folder_path, filename)
                # Abrimos el excel con la ruta creada anteriormente
                workbook = openpyxl.load_workbook(file_path)
                # Generamos una lista con los nombres de todas las hojas del archivo .xlxs
                sheet_names = workbook.sheetnames

                # En caso de que el nombre de la hoja se encuentre en la lista generada anteriormente la seleccionamos
                if sheet in sheet_names:
                    sheet_choice = workbook[sheet]
                    
                    # creamos una lista donde almacenaremos todos los datos extraidos de la hoja
                    data = []

                    # iteramos el rango de filas seleccionadas por el usuario y guardamos los datos de estas filas en la lista data
                    for row in sheet_choice.iter_rows(min_row = row_min, max_row=row_max, values_only=True):
                        data.append(row)

                    # añadimos estos datos a dataset     
                    dataset.extend(data)
                
                #Cerramos el excel
                workbook.close()
        
        # Convertimos todo lo almacenado del año en un dataframe
        df = pd.DataFrame(dataset)
        
        # Guardar df en formato pickle
        df.to_pickle(os.path.join(self.current_directory,"data","dataframes",str(name_df + ".pkl")))

    # Create excel_inflation folder
    def creation_folder(self):

        #Extraemos directorio actual y en caso de no tener una carpeta donde alojar los .xlsx la creamos
        self.excel_inflation_folder = self.current_directory + "/data/excel_inflation"
        

        # Create folder using os
        try:
            os.makedirs(self.excel_inflation_folder)
            print(f"Carpeta '{self.excel_inflation_folder}' creada correctamente.")
        except OSError as e:
            print(f"Error al crear la carpeta: {e}")

    # Check if there are data in the excel_infaltion folder for each month and year
    def data_folder_check(self):
        #Verificamos si existe algún archivo en la carpeta de destino, en caso negativo descargamos los .xlsx
        
        files_in_folder = os.listdir(self.excel_inflation_folder)
        
        if not files_in_folder:
            for año in self.años:
                for mes in self.meses:
                    url = "https://www.dane.gov.co/files/investigaciones/boletines/ipc/anexo_ipc_"
                    url = url + mes + año + ".xlsx"
                    response = requests.get(url)

                    with open(self.excel_inflation_folder + "/inflacion" + mes + año + ".xlsx" ,"wb") as file:
                        file.write(response.content)
        else:
            print("La carpeta contiene archivos:")

    # Download the data from the url
    def download_sheets(self):        

        # Check if the folder excel_inflation exists
    
        if not os.path.exists(self.excel_inflation_folder) :
            self.creation_folder()
        

        self.data_folder_check()


    '''

    # descargados todos los exceles debemos abrirlos, extraer los datos y pegarlos en un nuevo dataset añadiendo la columna data
    

    # c) osea que al correr este .py me descargue los archivos del DANE y despues los abra y me genere el xlsx que estoy necesitando, que ese sea el output
    # d) y luego que eliminé los archivos descargados, es decir una descarga temporal, porque no me interesa tenerlos en mi repositorio
    '''


