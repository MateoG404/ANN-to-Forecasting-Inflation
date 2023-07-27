import os
import pandas as pd
import requests
import openpyxl 


# Función que extrae los datos de determinado rango de filas y columnas de todos los exceles de una carpeta y da como output un dataframe

def extract_data(folder_path, row_min, row_max, column_min, column_max, year, sheet):
    
    dataset = []

    # recorremos todos los archivos dentro de la ruta indicada
    for filename in os.listdir(folder_path):
        # seleccionamos los archivos que terminen con el año seleccionado y que sean de tipo .xlxs
        if filename.endswith(str(year)+'.xlsx'):

            # Creamos la ruta del archivo uniendo la ruta del directorio y el nombre del archivo    
            file_path = os.path.join(folder_path, filename)
            # Abrimos el excel con la ruta creada anteriormente
            workbook = openpyxl.load_workbook(file_path)
            # Generamos una lista con los nombres de todas las hojas del archivo .xlxs
            sheet_names = workbook.sheetnames

            # En caso de que el nombre de la hoja se encuentre en la lista generada anteriormente la seleccionamos
            if sheet in sheet_names:
                sheet_choice = workbook[sheet]
                
                # creamos una lista donde almacenaremos todos los datos extraidos de la hoja
                data = []

                # iteramos el rango de filas seleccionadas por el usuario y guardamos los datos de estas filas en la lista data
                for row in sheet_choice.iter_rows(min_row = row_min, max_row=row_max, values_only=True):
                    # sliceamos solo las columnas de nuestro interes
                    sliced_row = row[column_min - 1: column_max] 
                    data.append(sliced_row)

                # añadimos estos datos a dataset     
                dataset.extend(data)
            
            #Cerramos el excel
            workbook.close()
    
    # Convertimos todo lo almacenado del año en un dataframe
    df = pd.DataFrame(dataset)

    #Añadimos una columna de fecha a ese dataframe
    date = []

    for mes in range(1,13):
        for n in range(row_max - row_min):
           date.append(str(mes) + "-" + str(year))

    df["date"] = date

    #devolvemos el dataframe
    return df

#Declaramos las variables necesarias para generar la URL e iterar para descargar los .xlxs

url = "https://www.dane.gov.co/files/investigaciones/boletines/ipc/anexo_ipc_"

meses = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
años = ["19", "20", "21", "22"]

#Extraemos directorio actual y en caso de no tener una carpeta donde alojar los .xlsx la creamos

current_directory = os.getcwd()
current_directory = current_directory.replace( "\\" , "/")
excel_inflation_folder = current_directory + "/excel_inflation"

if not os.path.isdir(excel_inflation_folder):
    os.makedirs(excel_inflation_folder)
    print("La carpeta fue creada exitosamente")
else:
    print("la carpeta ya existe")

#Verificamos si existe algún archivo en la carpeta de destino, en caso negativo descargamos los .xlsx

files_in_folder = os.listdir(excel_inflation_folder)

if not files_in_folder:
    for año in años:
        for mes in meses:
            url = "https://www.dane.gov.co/files/investigaciones/boletines/ipc/anexo_ipc_"
            url = url + mes + año + ".xlsx"
            response = requests.get(url)

            with open(excel_inflation_folder + "/inflacion" + mes + año + ".xlsx" ,"wb") as file:
                file.write(response.content)
else:
    print("La carpeta contiene archivos:")


# Corremos la función los años necesarios

df_2022 = extract_data(excel_inflation_folder, 10,197,1,9,22,8)
df_2021 = extract_data(excel_inflation_folder, 10,197,1,9,21,8)
df_2020 = extract_data(excel_inflation_folder, 10,197,1,9,20,8)
df_2019 = extract_data(excel_inflation_folder, 10,197,4,12,19,7)

print(df_2019)
print(df_2021)
print(df_2022)
print(df_2023)

# descargados todos los exceles debemos abrirlos, extraer los datos y pegarlos en un nuevo dataset añadiendo la columna data
 



# c) osea que al correr este .py me descargue los archivos del DANE y despues los abra y me genere el xlsx que estoy necesitando, que ese sea el output
# d) y luego que eliminé los archivos descargados, es decir una descarga temporal, porque no me interesa tenerlos en mi repositorio



